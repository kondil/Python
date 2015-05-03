#!/usr/bin/env python

# This is a small tutorial in Openpyxl.

# Let's read from a current excel file.
# - For this we will need to load 'load_workbook' from openpyxl.

from openpyxl import load_workbook

wb = load_workbook(filename='Nov.xlsx', read_only=True)
ws = wb['Nov.csv']

print ws.rows

# This will print the help.
help(ws.rows)


