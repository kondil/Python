#!/usr/bin/env python
# This is a tutorial on how to create an xlsx Workbook using the Openpyxl library.
# All infos retrieved from :: 
# https://openpyxl.readthedocs.org/en/latest/tutorial.html#create-a-workbook
from openpyxl import Workbook, styles

# Let's create a Workbook.
wb = Workbook()

# Let's grab the active Worksheet.
ws = wb.active
ws['A1'] = 42
ws.append([1,2,3])
ws = wb.create_sheet(0, 'tmp')
wb.get_sheet_names()

font = styles.Font(bold=True)

for row in range(ord('A'), ord('A')+15):
	for col in range(1,38):
		cell_pos = chr(row)+str(col)	# print cell_pos
		ws.cell(coordinate=cell_pos, value=cell_pos)
		ws[cell_pos].font = font
		# This one did not work for me :: ws.cell(row=row, column=col, value=cell_pos)
		# This is one is not working :: ws.cell(cell_pos, cell_pos)


wb.save("tmp.xlsx")

