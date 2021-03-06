#!/usr/bin/env python
# This is a tutorial on how to read an xlsx Workbook using the Openpyxl library.
# Let's read from a current excel file.
# - For this we will need to load 'load_workbook' from openpyxl.
from openpyxl import load_workbook

wb = load_workbook(filename='tmp.xlsx', read_only=True)

# In order to discover more info on the module, classes and objects we have loaded 
# we can use the commands dir() and help() like the next examples.
# dir(wb)
# help(wb)
# help(wb.get_active_sheet)


# ws = wb['Nov.csv']
# The following command will get us the active sheet of the workbook.
ws = wb.get_active_sheet()

# print ws.rows
# print ws.dimensions
# print ws.max_column
# print ws.max_row

# # To scan all cells in the Worksheet ::
# for row in ws.rows:
# 	print "=======================================\nNext Entry\n======================================="
# 	for cell in row:
# 		print cell.value
# 	print "=======================================\n"

# To scan a specific column of the Worksheet ::
for row in range(5, ws.max_row):
	if '1' in ws.columns[2][row].value:
		print row, ws.columns[2][row].value

